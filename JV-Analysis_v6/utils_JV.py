"""
Utility Functions Module
Contains Excel export, file operations, and other utility functions.
Extracted from main.py for better organization.
"""

__author__ = "Edgar Nandayapa"
__institution__ = "Helmholtz-Zentrum Berlin"
__created__ = "August 2025"

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd
import os


def save_full_data_frame(data):
    """
    Create and return an Excel workbook with the full dataframe.
    Simplified version that just creates a workbook without saving to file.
    """
    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)  # Remove the default sheet
    
    # Add main data sheet
    ws = wb.create_sheet(title='All_data')
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)
    
    return wb


def save_combined_excel_data(path, wb, data, filtered_info, var_x, name_y, var_y, other_df):
    """Save combined data to Excel workbook with multiple sheets"""
    trash, filters = filtered_info
    
    # Create sheet name based on variables
    sheet_title = f"{var_y}-by-{var_x}"

    # Check if the sheet already exists and remove it
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    # Insert header
    ws.append([f"Contents of boxplot for {var_y} by {var_x}"])
    ws.append([])  # Empty row

    # Process and append main data
    combined_data = data.copy()
    combined_data['_index'] = combined_data.groupby(var_x).cumcount()
    pivot_table = combined_data.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

    for r in dataframe_to_rows(pivot_table, index=True, header=True):
        ws.append(r)

    # Add statistical summary
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="Statistical summary")
    ws.append([])

    for r in dataframe_to_rows(other_df.T, index=True, header=True):
        ws.append(r)

    # Add filtered data section
    next_row = ws.max_row + 3
    ws.cell(row=next_row, column=1, value="This is the filtered data")
    ws.append([])

    if not trash.empty:
        combined_trash = trash.copy()
        combined_trash['_index'] = combined_trash.groupby(var_x).cumcount()
        pivot_table_trash = combined_trash.pivot_table(index='_index', columns=var_x, values=name_y, aggfunc="mean")

        for r in dataframe_to_rows(pivot_table_trash, index=True, header=True):
            ws.append(r)

    # Add filter information
    next_row = ws.max_row + 3
    filter_words = ["Only data within these limits is shown:"] + filters
    for cc, strings in enumerate(filter_words):
        ws.cell(row=next_row + cc, column=1, value=strings)

    return wb


def is_running_in_jupyter():
    """Check if code is running in Jupyter notebook"""
    try:
        from IPython.core.getipython import get_ipython
        return get_ipython() is not None
    except (ImportError, AttributeError):
        return False


def create_new_results_folder(path):
    """Create a results folder if it doesn't exist"""
    folder_path = os.path.join(path, 'Results')
    try:
        os.makedirs(folder_path, exist_ok=True)
    except Exception as e:
        print(f"Warning: Could not create results folder: {e}")
        return path
    return folder_path


def clean_filename(filename):
    """Clean filename for safe saving"""
    import re
    # Remove invalid characters for filenames
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    return filename


def generate_detailed_export_excel(export_df, filtered_info=None):
    """
    Generate Excel workbook with detailed export data.
    
    Creates multiple sheets:
    - Rohdaten: All pixel-level data with filter info and champion/median markings
    - Zusammenfassung: Summary statistics per variation
    - Filter-Log: Detailed log of excluded data with reasons
    
    Args:
        export_df: DataFrame from export_detailed_pixel_data()
        filtered_info: Tuple of (trash_df, filter_reasons_list) for additional context
    
    Returns:
        openpyxl.Workbook object ready for saving
    """
    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)  # Remove default sheet
    
    # ==================== Sheet 1: Rohdaten ====================
    ws_raw = wb.create_sheet(title='Rohdaten')
    
    # Add header
    ws_raw.append(['DETAILED PIXEL-LEVEL DATA'])
    ws_raw.append(['All measurements with filter status, filter reasons, and champion/median markings'])
    ws_raw.append([])
    
    # Add data
    for r in dataframe_to_rows(export_df, index=False, header=True):
        ws_raw.append(r)
    
    # Format header row
    header_fill = PatternFill(start_color="1F3964", end_color="1F3964", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws_raw[4]:  # Row 4 is the header
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
    
    # Auto-adjust column widths
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header rows
    ws_raw.freeze_panes = 'A5'
    
    # ==================== Sheet 2: Zusammenfassung ====================
    ws_summary = wb.create_sheet(title='Zusammenfassung')
    
    ws_summary.append(['SUMMARY STATISTICS PER VARIATION'])
    ws_summary.append(['Champion and Median values for each variation'])
    ws_summary.append([])
    
    # Group by variation/identifier
    groupby_col = 'identifier' if 'identifier' in export_df.columns else 'sample'
    
    summary_rows = []
    summary_rows.append(['Variation', 'Total Pixels', 'Included', 'Excluded', 
                        'PCE(%) Mean', 'PCE(%) Min', 'PCE(%) Max',
                        'Voc(V) Mean', 'Jsc(mA/cm2) Mean', 'FF(%) Mean'])
    
    for variation in export_df[groupby_col].unique():
        var_data = export_df[export_df[groupby_col] == variation]
        included = len(var_data[var_data['filter_status'] == 'Included'])
        excluded = len(var_data[var_data['filter_status'] == 'Excluded'])
        
        # Get only included data for statistics
        included_data = var_data[var_data['filter_status'] == 'Included']
        
        pce_mean = included_data['PCE(%)'].mean() if 'PCE(%)' in included_data.columns else 0
        pce_min = included_data['PCE(%)'].min() if 'PCE(%)' in included_data.columns else 0
        pce_max = included_data['PCE(%)'].max() if 'PCE(%)' in included_data.columns else 0
        voc_mean = included_data['Voc(V)'].mean() if 'Voc(V)' in included_data.columns else 0
        jsc_mean = included_data['Jsc(mA/cm2)'].mean() if 'Jsc(mA/cm2)' in included_data.columns else 0
        ff_mean = included_data['FF(%)'].mean() if 'FF(%)' in included_data.columns else 0
        
        summary_rows.append([
            str(variation),
            len(var_data),
            included,
            excluded,
            round(pce_mean, 2),
            round(pce_min, 2),
            round(pce_max, 2),
            round(voc_mean, 3),
            round(jsc_mean, 2),
            round(ff_mean, 2)
        ])
    
    for row in summary_rows:
        ws_summary.append(row)
    
    # Format header row
    for cell in ws_summary[4]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # ==================== Sheet 3: Filter-Log ====================
    ws_log = wb.create_sheet(title='Filter-Log')
    
    ws_log.append(['FILTER LOG - EXCLUDED DATA'])
    ws_log.append(['Details of all measurements that were filtered out and reasons'])
    ws_log.append([])
    
    excluded_data = export_df[export_df['filter_status'] == 'Excluded'].copy()
    
    if not excluded_data.empty:
        # Select relevant columns for the log
        log_cols = ['sample', 'cell', 'px_number', 'cycle_number', 
                   'Voc(V)', 'Jsc(mA/cm2)', 'FF(%)', 'PCE(%)',
                   'filter_reason', 'identifier']
        log_cols = [col for col in log_cols if col in excluded_data.columns]
        
        log_df = excluded_data[log_cols].copy()
        
        for r in dataframe_to_rows(log_df, index=False, header=True):
            ws_log.append(r)
        
        # Format header row
        for cell in ws_log[4]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
        
        # Auto-adjust column widths
        for column in ws_log.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_log.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header rows
        ws_log.freeze_panes = 'A5'
    else:
        ws_log.append(['No excluded data - all measurements passed filters'])
    
    # ==================== Sheet 4: Filter-Gründe Summary ====================
    ws_reasons = wb.create_sheet(title='Filter-Gruende')
    
    ws_reasons.append(['FILTER REASONS SUMMARY'])
    ws_reasons.append(['Count of measurements excluded for each reason'])
    ws_reasons.append([])
    
    if not excluded_data.empty and 'filter_reason' in excluded_data.columns:
        reason_counts = excluded_data['filter_reason'].value_counts()
        
        ws_reasons.append(['Reason', 'Count', 'Percentage'])
        total_excluded = len(excluded_data)
        
        for reason, count in reason_counts.items():
            percentage = round((count / total_excluded) * 100, 1) if total_excluded > 0 else 0
            ws_reasons.append([str(reason), count, f"{percentage}%"])
    else:
        ws_reasons.append(['No filter reasons recorded'])
    
    # Format header row
    for cell in ws_reasons[4]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
    
    # Auto-adjust column widths
    for column in ws_reasons.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws_reasons.column_dimensions[column_letter].width = adjusted_width
    
    return wb