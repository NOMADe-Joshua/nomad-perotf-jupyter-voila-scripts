"""
Utility Functions Module
Contains Excel export, file operations, and other utility functions.
Copied from JV-Analysis for UVVis Analyzer.
"""

__author__ = "Joshua Damm"
__institution__ = "KIT"
__created__ = "December 2025"

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import os


def save_full_data_frame(data):
    """
    Create and return an Excel workbook with the full dataframe.
    Simplified version that just creates a workbook without saving to file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Add main data sheet
    ws = wb.create_sheet(title='All_data')
    for r in dataframe_to_rows(data, index=True, header=True):
        ws.append(r)
    
    return wb


def is_running_in_jupyter():
    """Check if code is running in Jupyter notebook"""
    try:
        from IPython import get_ipython
        return get_ipython() is not None
    except ImportError:
        return False


def create_new_results_folder(path):
    """Create a results folder if it doesn't exist"""
    folder_path = os.path.join(path, 'Results')
    try:
        os.makedirs(folder_path, exist_ok=True)
    except Exception as e:
        print(f"Warning: Could not create results folder: {e}")
        return path
    return folder_path


def clean_filename(filename):
    """Clean filename for safe saving"""
    import re
    # Remove invalid characters for filenames
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    return filename
