from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

TABLEAU_COLORS = {
    'tab:blue': '1F77B4',
    'tab:orange': 'FF7F0E',
    'tab:green': '2CA02C',
    'tab:red': 'D62728',
    'tab:purple': '9467BD',
    'tab:brown': '8C564B',
    'tab:pink': 'E377C2',
    'tab:gray': '7F7F7F',
    'tab:olive': 'BCBD22',
    'tab:cyan': '17BECF',
}
colors = list(TABLEAU_COLORS.values())


def lighten_color(hex_color, factor=0.50):
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f'{r:02x}{g:02x}{b:02x}'.upper()


def add_experiment_sheet(workbook, process_sequence, is_testing=False):
    ws = workbook.active
    ws.title = 'Experiment Data'
    start_col = 1
    incremental_number = 0

    def make_label(label, test_val=None):
        if is_testing:
            if test_val is not None:
                return (label, test_val)
            else:
                return (label, '')
        else:
            return label

    def generate_steps_for_process(process_name, config):
        """
        This method constructs steps for each process. If is_testing=True,
        call make_label("Step Name", <test_value>) to pass a custom test value.
        Otherwise, just pass "Step Name".
        """

        if process_name == 'Experiment Info':
            return [
                make_label('Date', 'YYYYMMDD'),
                make_label('Project_Name', 'NaMe'),
                make_label('Batch', '0'),
                make_label('Subbatch', '0'),
                make_label('Sample', '0'),
                make_label('Nomad ID', ''),
                make_label('Variation', 'readable variation'),
                make_label('Sample dimension', '16x16'), 
                make_label('Sample area [cm^2]', '0.105'),
                make_label('Number of pixels', '4'),
                make_label('Pixel area', '0.105'),
                make_label('Number of junctions', '1'),
                make_label('Substrate material', 'Glass'),
                make_label('Substrate conductive layer', 'ITO'),
                make_label('Bottom Cell Name', ''),
                make_label('Notes', ''),
            ]

        if process_name == 'Multijunction Info':
            return [
                make_label('Recombination Layer', ''),
                make_label('Notes', ''),
            ]

        if process_name == 'Cleaning O2-Plasma' or process_name == 'Cleaning UV-Ozone':
            steps = []
            for i in range(1, config.get('solvents', 1) + 1):
                steps.extend([
                    make_label(f'Solvent {i}', ''),
                    make_label(f'Time {i} [s]', '900'),
                    make_label(f'Temperature {i} [°C]', '40'),
                ])

            if process_name == 'Cleaning O2-Plasma':
                steps.extend([
                    make_label('Gas-Plasma Gas', ''),
                    make_label('Gas-Plasma Time [s]', ''),
                    make_label('Gas-Plasma Power [W]', ''),
                ])
            if process_name == 'Cleaning UV-Ozone':
                steps.append(make_label('UV-Ozone Time [s]', '900'))
            return steps

        if process_name in ['Spin Coating', 'Dip Coating', 'Slot Die Coating', 'Inkjet Printing']:
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', ''),
            ]

            # Add solvent steps
            for i in range(1, config.get('solvents', 1) + 1):
                steps.extend([
                    make_label(f'Solvent {i} name', ''),
                    make_label(f'Solvent {i} volume [uL]', ''),
                ])

            # Add solute steps
            for i in range(1, config.get('solutes', 1) + 1):
                steps.extend([
                    make_label(f'Solute {i} type', ''),
                    make_label(f'Solute {i} Concentration [mM]', ''),
                ])

            # Add process-specific steps
            if process_name == 'Spin Coating':
                steps.extend([
                    make_label('Solution volume [uL]', ''),
                    make_label('Spin Delay [s]', '')
                ])

                if config.get('spinsteps', 1) == 1:
                    steps.extend([
                        make_label('Rotation speed [rpm]', ''),
                        make_label('Rotation time [s]', ''),
                        make_label('Acceleration [rpm/s]', ''),
                    ])
                else:
                    for i in range(1, config.get('spinsteps', 1) + 1):
                        steps.extend([
                            make_label(f'Rotation speed {i} [rpm]', ''),
                            make_label(f'Rotation time {i} [s]', ''),
                            make_label(f'Acceleration {i} [rpm/s]', ''),
                        ])

                if config.get('antisolvent', False):
                    steps.extend([
                        make_label('Anti solvent name', ''),
                        make_label('Anti solvent volume [ml]', ''),
                        make_label('Anti solvent dropping time [s]', ''),
                        make_label('Anti solvent dropping speed [uL/s]', ''),
                        make_label('Anti solvent dropping heigt [mm]', ''),
                    ])

                if config.get('gasquenching', False):
                    steps.extend([
                        make_label('Gas', ''),
                        make_label('Gas quenching start time [s]', ''),
                        make_label('Gas quenching duration [s]', ''),
                        make_label('Gas quenching flow rate [ml/s]', ''),
                        make_label('Gas quenching pressure [bar]', ''),
                        make_label('Gas quenching velocity [m/s]', ''),
                        make_label('Gas quenching height [mm]', ''),
                        make_label('Nozzle shape', ''),
                        make_label('Nozzle size [mm²]', ''),
                    ])

                if config.get('vacuumquenching', False):
                    steps.extend([
                        make_label('Vacuum quenching start time [s]', ''),
                        make_label('Vacuum quenching duration [s]', ''),
                        make_label('Vacuum quenching pressure [bar]', ''),
                    ])

            elif process_name == 'Slot Die Coating':
                steps.extend([
                    make_label('Coating run', ''),
                    make_label('Solution volume [um]', ''),
                    make_label('Flow rate [uL/min]', ''),
                    make_label('Head gap [mm]', ''),
                    make_label('Speed [mm/s]', ''),
                    make_label('Air knife angle [°]', ''),
                    make_label('Air knife gap [cm]', ''),
                    make_label('Bead volume [mm/s]', ''),
                    make_label('Drying speed [cm/min]', ''),
                    make_label('Drying gas temperature [°]', ''),
                    make_label('Heat transfer coefficient [W m^-2 K^-1]', ''),
                    make_label('Coated area [mm²]', ''),
                ])

            elif process_name == 'Dip Coating':
                steps.append(make_label('Dipping duration [s]', ''))

            elif process_name == 'Inkjet Printing':
                steps.extend([
                    make_label('Printhead name', ''),
                    make_label('Printing run', ''),
                    make_label('Number of active nozzles', ''),
                    make_label('Droplet density [dpi]', ''),
                    make_label('Quality factor', ''),
                    make_label('Step size', ''),
                    make_label('Printing direction', ''),
                    make_label('Printed area [mm²]', ''),
                    make_label('Droplet per second [1/s]', ''),
                    make_label('Droplet volume [pL]', ''),
                    make_label('Dropping Height [mm]', ''),
                    make_label('Ink reservoir pressure [mbar]', ''),
                    make_label('Table temperature [°C]', ''),
                    make_label('Nozzle temperature [°C]', ''),
                    make_label('Room temperature [°C]', ''),
                    make_label('rel. humidity [%]', ''),
                ])

                pixORnotion = config.get('pixORnotion', 'Pixdro')
                if pixORnotion == 'Pixdro':
                    steps.append(make_label('Wf Number of Pulses', ''))
                    for N_Pulse in range(1, config.get('Wf Number of Pulses', 1) + 1):
                        steps.extend([
                            make_label(f'Wf Level {N_Pulse}[V]', ''),
                            make_label(f'Wf Rise {N_Pulse}[V/us]', ''),
                            make_label(f'Wf Width {N_Pulse}[us]', ''),
                            make_label(f'Wf Fall {N_Pulse}[V/us]', ''),
                            make_label(f'Wf Space {N_Pulse}[us]', ''),
                        ])

                if pixORnotion == 'Notion':
                    steps.extend([
                        make_label('Wf Number of Pulses', ''),
                        make_label('Wf Delay Time [us]', ''),
                        make_label('Wf Rise Time [us]', ''),
                        make_label('Wf Hold Time [us]', ''),
                        make_label('Wf Fall Time [us]', ''),
                        make_label('Wf Relax Time [us]', ''),
                        make_label('Wf Voltage [V]', ''),
                        make_label('Wf Number Greylevels', ''),
                        make_label('Wf Grey Level 0 Use Pulse [1/0]', ''),
                        make_label('Wf Grey Level 1 Use Pulse [1/0]', ''),
                    ])

                if config.get('gasquenching', False):
                    steps.extend([
                        make_label('Gas', ''),
                        make_label('Gas quenching start time [s]', ''),
                        make_label('Gas quenching duration [s]', ''),
                        make_label('Gas quenching flow rate [ml/s]', ''),
                        make_label('Gas quenching pressure [bar]', ''),
                        make_label('Gas quenching velocity [m/s]', ''),
                        make_label('Gas quenching height [mm]', ''),
                        make_label('Nozzle shape', ''),
                        make_label('Nozzle size [mm²]', ''),
                    ])

                if config.get('vacuumquenching', False):
                    steps.extend([
                        make_label('Vacuum quenching start time [s]', ''),
                        make_label('Vacuum quenching duration [s]', ''),
                        make_label('Vacuum quenching pressure [bar]', ''),
                    ])

            # Add annealing steps for all coating processes
            steps.extend([
                make_label('Annealing time [min]', ''),
                make_label('Annealing temperature [°C]', ''),
                make_label('Annealing athmosphere', ''),
                make_label('Notes', ''),
            ])

            return steps

        # PVD Processes
        if process_name == 'Evaporation':
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', ''),
                make_label('Organic', ''),
                make_label('Base pressure [bar]', ''),
                make_label('Pressure start [bar]', ''),
                make_label('Pressure end [bar]', ''),
                make_label('Source temperature start[°C]', ''),
                make_label('Source temperature end[°C]', ''),
                make_label('Substrate temperature [°C]', ''),
                make_label('Thickness [nm]', ''),
                make_label('Rate [angstrom/s]', ''),
                make_label('Power [%]', ''),
                make_label('Tooling factor', ''),
                make_label('Notes', ''),
            ]
            return steps

        if process_name == 'Co-Evaporation' or process_name == 'Seq-Evaporation':
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', ''),
            ]
            for i in range(1, config.get('materials', 2) + 1):
                steps.extend([
                    make_label(f'Material name {i}', ''),
                    make_label(f'Base pressure {i} [bar]', ''),
                    make_label(f'Pressure start {i} [bar]', ''),
                    make_label(f'Pressure end {i} [bar]', ''),
                    make_label(f'Source temperature start {i}[°C]', ''),
                    make_label(f'Source temperature end {i}[°C]', ''),
                    make_label(f'Substrate temperature {i} [°C]', ''),
                    make_label(f'Thickness {i} [nm]', ''),
                    make_label(f'Rate {i} [angstrom/s]', ''),
                    make_label(f'Tooling factor {i}', '')
                ])
            steps.append(make_label('Notes', ''))
            return steps

        if process_name == 'Close Space Sublimation':
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', 'Solvent GB'),
                make_label('Organic', ''),
                make_label('Process pressure [bar]', '4'),
                make_label('Source temperature [°C]', ''),
                make_label('Substrate temperature [°C]', ''),
                make_label('Material state', ''),
                make_label('Substrate source distance [mm]', '4'),
                make_label('Thickness [nm]', ''),
                make_label('Deposition Time [s]', ''),
                make_label('Carrier gas', 'no'),
                make_label('Notes', ''),
            ]
            return steps

        if process_name == 'Lamination':
            steps = [
                make_label('Interface', ''),
                make_label('Tool/GB name', ''),
                make_label('Temperature during process[°C]', ''),
                make_label('Temperature at pressure relief [°C]', ''),
                make_label('Pressure [MPa]', ''),
                make_label('Force [N]', ''),
                make_label('Time lamination [s]', ''),
                make_label('Heat up time [s]', ''),
                make_label('Cool down time [s]', ''),
                make_label('Total time [s]', ''),
                make_label('Athmosphere in chamber', ''),
                make_label('Humidity [%%rel]', ''),
                make_label('Stamp 1 Material', ''),
                make_label('Stamp 1 Thickness [mm]', ''),
                make_label('Stamp 1 Area [mm^2]', ''),
                make_label('Stamp 2 Material', ''),
                make_label('Stamp 2 Thickness [mm]', ''),
                make_label('Stamp 2 Area [mm^2]', ''),
                make_label('Homogeniously pressed [1/0]', ''),
                make_label('Sucessful adhesion [1/0]', ''),
                make_label('Notes', ''),
            ]
            return steps

        if process_name == 'Sputtering':
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', ''),
                make_label('Gas', ''),
                make_label('Temperature [°C]', ''),
                make_label('Pressure [mbar]', ''),
                make_label('Deposition time [s]', ''),
                make_label('Burn in time [s]', ''),
                make_label('Power [W]', ''),
                make_label('Rotation rate [rpm]', ''),
                make_label('Thickness [nm]', ''),
                make_label('Gas flow rate [cm^3/min]', ''),
                make_label('Notes', ''),
            ]
            return steps

        if process_name == 'Laser Scribing':
            steps = [
                make_label('Laser wavelength [nm]', ''),
                make_label('Laser pulse time [ps]', ''),
                make_label('Laser pulse frequency [kHz]', ''),
                make_label('Speed [mm/s]', ''),
                make_label('Fluence [J/cm2]', ''),
                make_label('Power [%]', ''),
                make_label('Recipe file', ''),
            ]
            return steps

        if process_name == 'ALD':
            steps = [
                make_label('Material name', ''),
                make_label('Layer type', ''),
                make_label('Tool/GB name', ''),
                make_label('Source', ''),
                make_label('Thickness [nm]', ''),
                make_label('Temperature [°C]', ''),
                make_label('Rate [A/s]', ''),
                make_label('Time [s]', ''),
                make_label('Number of cycles', ''),
                make_label('Precursor 1', ''),
                make_label('Pulse duration 1 [s]', ''),
                make_label('Manifold temperature 1 [°C]', ''),
                make_label('Bottle temperature 1 [°C]', ''),
                make_label('Precursor 2 (Oxidizer/Reducer)', ''),
                make_label('Pulse duration 2 [s]', ''),
                make_label('Manifold temperature 2 [°C]', ''),
            ]
            return steps

        if process_name == 'Annealing':
            steps = [
                make_label('Annealing time [min]', ''),
                make_label('Annealing temperature [°C]', ''),
                make_label('Annealing athmosphere', ''),
                make_label('Relative humidity [%]', ''),
                make_label('Notes', ''),
            ]
            return steps

        if process_name == 'Generic Process':
            steps = [
                make_label('Name', ''),
                make_label('Notes', ''),
            ]
            return steps

        else:
            print(f"Warning: Process '{process_name}' not defined in generate_steps_for_process. Using default steps.")
            return [make_label('Undefined Process', '')]

    for process_data in process_sequence:
        process_name = process_data['process']
        custom_config = process_data.get('config', {})
        color_index = incremental_number % len(colors)
        cell_color = colors[color_index]
        steps = generate_steps_for_process(process_name, custom_config)
        step_count = len(steps)
        end_col = start_col + step_count - 1

        if process_name != 'Experiment Info':
            process_label = f'{incremental_number}: {process_name}'
        else:
            process_label = process_name

        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col)
        cell.value = process_label
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=cell_color,
                                end_color=cell_color, fill_type='solid')

        row2_color = lighten_color(cell_color)
        for i, step_item in enumerate(steps):
            col_index = start_col + i
            if isinstance(step_item, tuple):
                step_label, test_val = step_item
                cell = ws.cell(row=2, column=col_index)
                cell.value = step_label
                cell.fill = PatternFill(start_color=row2_color,
                                        end_color=row2_color, fill_type='solid')
                if is_testing:
                    ws.cell(row=3, column=col_index, value=test_val)
        start_col = end_col + 1
        incremental_number += 1

    # Example: Apply a custom formula for the "Nomad ID" column (example only)
    for row in range(3, 4):
        nomad_id_formula = f'=CONCATENATE("HZB_",B{row},"_",C{row},"_",D{row},"_",E{row})'
        ws[f'F{row}'].value = nomad_id_formula

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and isinstance(cell.value, str):
                max_length = max(max_length, len(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2
